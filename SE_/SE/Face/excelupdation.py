import sys
#from colorama import Fore, init, Back, Style
import openpyxl
import re
import time
from datetime import date, datetime
from .emailing import sendmail

def updateexcel(name):
    flag=0
    path = "/Users/shashwateshtripathi/Desktop/SE/SE/Face/attendance.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    max_column = sheet_obj.max_column
    max_row = sheet_obj.max_row
    percentage=0
    if sheet_obj.cell(row=1,column=max_column).value !=  str(date.today()):
        sheet_obj.cell(row=1,column=max_column+1).value = str(date.today())
        print(str(date.today()))
        for j in range(2,max_row+1):
            sheet_obj.cell(row=j,column=max_column+1).value = 0
    for j in range(2,max_row+1):
        max_column=sheet_obj.max_column
        if sheet_obj.cell(row=j,column=1).value == name.upper():
            sheet_obj.cell(row=j,column=max_column).value = 1
            count=0
            for i in range(4,max_column+1):
                if sheet_obj.cell(row=j,column=i).value==1:
                    count+=1
            percentage=float(count/(max_column-3))*100
            flag=1
            break
    if flag==0:
        return -1
    wb_obj.save("/Users/shashwateshtripathi/Desktop/SE/SE/Face/attendance.xlsx")
    return percentage

def getstudentdetails():
    path = "/Users/shashwateshtripathi/Desktop/SE/SE/Face/attendance.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj=wb_obj.active
    max_column = sheet_obj.max_column
    max_row = sheet_obj.max_row
    stud_detail = {}
    lolwa=0
    for j in range(2,max_row+1):
        l=[]
        count=0
        for i in range(4,max_column+1):
            if sheet_obj.cell(row=j,column=i).value==1:
                count+=1
        percentage=float(count/(max_column-3))*100
        l.append(sheet_obj.cell(row=j,column=1).value)
        l.append(sheet_obj.cell(row=j,column=3).value)
        l.append(percentage)
        stud_detail[sheet_obj.cell(row=j,column=2).value]=l
        if percentage<50:
            lolwa=lolwa+1
    return stud_detail,lolwa

def newstudentexcel(name,roll,email):
    path = "/Users/shashwateshtripathi/Desktop/SE/SE/Face/attendance.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    max_column = sheet_obj.max_column
    max_row = sheet_obj.max_row
    sheet_obj.cell(row=max_row+1,column =1).value = name.upper()
    sheet_obj.cell(row=max_row+1,column =2).value = roll
    sheet_obj.cell(row=max_row+1,column =3).value = email.lower()
    max_column = sheet_obj.max_column
    if max_column>2:
        for j in range(4,max_column+1):
            sheet_obj.cell(row=max_row+1,column=j).value=0
        
    wb_obj.save(path)

def mailexcel(percent):
    path = "/Users/shashwateshtripathi/Desktop/SE/SE/Face/attendance.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj=wb_obj.active
    max_column = sheet_obj.max_column
    max_row = sheet_obj.max_row
    stud_detail = []
    lolwa=0
    for j in range(2,max_row+1):
        count=0
        for i in range(4,max_column+1):
            if sheet_obj.cell(row=j,column=i).value==1:
                count+=1
        percentage=float(count/(max_column-3))*100
        if percentage<=int(percent):
            sendmail(str(sheet_obj.cell(row=j,column=3).value))